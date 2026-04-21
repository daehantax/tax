#!/usr/bin/env python3
"""조특법.xlsx '혜택내용' 시트의 빈 칸을 지원 데이터로 자동 보강.

채우는 칸:
  - agri_tax (농특세)         → 농특세 통합DB 기준
  - min_tax (최저한세)         → 132조 법인세/소득세/제외 데이터 기준
  - exclusion_codes (중복지원) → 127조 매트릭스 기준 (G2/G4/G5 그룹 추정)
  - deemed_depr (감가상각의제) → 세액감면/공제 패턴 기반 추정 (검토 필요)

출력:
  data/조특법_자동보강_검토용.xlsx
    - 자동 채운 셀: 노란색 배경 + 굵은 글씨
    - 추정 셀(deemed_depr): 주황색 배경 + ★ 마크
    - 사용자 검증 후 원본 조특법.xlsx 에 반영하시기 바람.
"""
from __future__ import annotations
import json
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
SRC_XLSX = DATA / "조특법.xlsx"
OUT_XLSX = DATA / "조특법_자동보강_검토용.xlsx"

FILL_AUTO = PatternFill("solid", fgColor="FEF3C7")    # 노란색 - 데이터 기반
FILL_GUESS = PatternFill("solid", fgColor="FED7AA")   # 주황색 - 추정
FONT_BOLD = Font(bold=True, color="92400E")

# ─── article_key 정규화 (변환 스크립트와 동일 로직) ─────────────────────────
CIRCLED = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"

def circled_to_num(s: str) -> str:
    for i, ch in enumerate(CIRCLED, 1):
        s = s.replace(ch, f"({i})")
    return s

def normalize_article_key(s) -> str:
    if s is None: return ""
    t = str(s).strip()
    if not t: return ""
    t = circled_to_num(t)
    t = t.replace("조특법", "").replace("제", "").replace("§", "").strip()
    hang = ""
    m = re.search(r"(?:\s+|제)(\d+)\s*항", t)
    if m:
        hang = "." + m.group(1)
        t = t[:m.start()] + t[m.end():]
    m = re.search(r"\((\d+)\)", t)
    if m:
        if not hang: hang = "." + m.group(1)
        t = t[:m.start()] + t[m.end():]
    m = re.match(r".*?(\d+)\s*[~\-]\s*\d+\s*항", t)
    if m and not hang:
        hang = "." + m.group(1)
        t = re.sub(r"\s*\d+\s*[~\-]\s*\d+\s*항.*$", "", t)
    t = t.replace("조", "").replace("의", "-").strip()
    t = re.sub(r"\s+", "", t)
    return t + hang


def load_json(name):
    return json.loads((DATA / name).read_text(encoding="utf-8"))


def main():
    detail = load_json("조특법_조문상세.json")
    matrix = load_json("조특법_중복배제매트릭스.json").get("pairs", [])

    # 127 매트릭스: article_key → [basis 리스트]
    key_to_basis = {}
    for p in matrix:
        for k in (p["a_key"], p["b_key"]):
            if not k: continue
            # 조 단위 키도 등록 (8-3.1, 8-3.3 모두 8-3 으로 매칭 가능하게)
            base = k.split(".")[0]
            for kk in {k, base}:
                key_to_basis.setdefault(kk, []).append(p["basis"])

    # basis (§127② 등) → G2/G3/G4/G5 매핑
    def basis_to_group(basis: str) -> str:
        # 원문자 변환
        for src, dst in zip(CIRCLED, "1234567890"):
            basis = basis.replace(src, dst)
        m = re.search(r"127\D*(\d+)", basis)
        if not m: return ""
        n = m.group(1)
        return {"2":"G2","3":"G3","4":"G4","5":"G5"}.get(n, "")

    # ───────── xlsx 보강 ─────────
    wb = openpyxl.load_workbook(SRC_XLSX)
    ws = wb["혜택내용"]

    # 컬럼: 1=id, 2=article, 3=title, 7=exclusion_codes, 8=agri_tax, 9=min_tax, 10=deemed_depr
    stats = {"agri_tax":0, "min_tax":0, "exclusion_codes":0, "deemed_depr":0, "skipped":0}

    for r in range(2, ws.max_row + 1):
        article = ws.cell(r, 2).value
        title = ws.cell(r, 3).value or ""
        if not (article or title):
            continue
        akey = normalize_article_key(article)
        akey_base = akey.split(".")[0] if akey else ""
        # detail 조회: 정확한 키 → 조 단위 키 순으로
        det = detail.get(akey) or detail.get(akey_base) or {}

        # ─ agri_tax (8) ─────────────────────────────────────────────
        cell8 = ws.cell(r, 8)
        if not (cell8.value and str(cell8.value).strip()):
            at = det.get("agri_tax") or {}
            status = at.get("status", "")
            if status:
                # "비과세" → "해당없음", "과세" → "해당"
                if "비과세" in status: val = "해당없음"
                elif "과세" in status: val = "해당"
                else: val = status
                basis = at.get("basis", "")
                cell8.value = val
                cell8.fill = FILL_AUTO
                cell8.font = FONT_BOLD
                if basis:
                    cell8.comment = Comment(f"근거: {basis}", "auto-fill")
                stats["agri_tax"] += 1

        # ─ min_tax (9) ──────────────────────────────────────────────
        cell9 = ws.cell(r, 9)
        if not (cell9.value and str(cell9.value).strip()):
            mt = det.get("min_tax_132") or {}
            applies = mt.get("applies") or []
            flag = mt.get("flag", "")
            if applies:
                # 제외 항목이면 해당없음
                kinds = {a.get("kind","") for a in applies}
                if kinds == {"제외"}:
                    val = "해당없음"
                else:
                    val = "해당"
                cell9.value = val
                cell9.fill = FILL_AUTO
                cell9.font = FONT_BOLD
                cmt_lines = [f"{a.get('kind','')}: {a.get('name','')} {a.get('note','')}".strip()
                             for a in applies]
                cell9.comment = Comment("\n".join(cmt_lines)[:1000], "auto-fill")
                stats["min_tax"] += 1
            elif flag:
                # flag 만 있는 경우 (예: "O", "O(100%연도제외)")
                val = "해당" if flag.startswith("O") else "해당없음"
                cell9.value = val
                cell9.fill = FILL_AUTO
                cell9.font = FONT_BOLD
                if "100%연도제외" in flag:
                    cell9.comment = Comment("100% 감면연도는 제외", "auto-fill")
                stats["min_tax"] += 1

        # ─ exclusion_codes (7) ─────────────────────────────────────
        cell7 = ws.cell(r, 7)
        if not (cell7.value and str(cell7.value).strip()):
            bases = key_to_basis.get(akey, []) + key_to_basis.get(akey_base, [])
            groups = sorted({basis_to_group(b) for b in bases if basis_to_group(b)})
            if groups:
                cell7.value = ", ".join(groups)
                cell7.fill = FILL_AUTO
                cell7.font = FONT_BOLD
                stats["exclusion_codes"] += 1
            else:
                cell7.value = "NONE"
                cell7.fill = FILL_AUTO
                cell7.font = FONT_BOLD
                stats["exclusion_codes"] += 1

        # ─ deemed_depr (10) — 추정 기반 ─────────────────────────────
        cell10 = ws.cell(r, 10)
        if not (cell10.value and str(cell10.value).strip()):
            t = title
            has_gameon = any(k in t for k in ["감면", "면제", "특례"])
            has_gongje = "공제" in t
            if has_gameon and not has_gongje:
                val = "해당"           # 세액감면 → 일반적으로 감가상각의제 적용
            elif has_gongje and not has_gameon:
                val = "해당없음"       # 세액공제 → 일반적으로 미적용
            elif has_gameon and has_gongje:
                val = "해당"           # 둘 다 있으면 안전하게 '해당'
            else:
                val = ""               # 판단 불가 → 비워둠
            if val:
                cell10.value = "★ " + val   # ★ 표시로 추정임을 강조
                cell10.fill = FILL_GUESS
                cell10.font = FONT_BOLD
                cell10.comment = Comment(
                    "★ 자동 추정 값입니다. 실제 적용 여부는 조문 검토 후 확정해 주세요.\n"
                    "기준: 세액감면 → '해당' / 세액공제 → '해당없음'", "auto-guess")
                stats["deemed_depr"] += 1

    # ─ 안내 시트 추가 ───────────────────────────────────────────────
    if "_자동보강_안내" in wb.sheetnames:
        wb.remove(wb["_자동보강_안내"])
    info = wb.create_sheet("_자동보강_안내", 0)  # 첫 시트로 배치
    info.column_dimensions["A"].width = 100
    rows = [
        ("📋 자동 보강 결과 — 검토용 파일",                          24, "FEF3C7"),
        ("",                                                         12, ""),
        (f"  ✓ agri_tax(농특세) 보강:        {stats['agri_tax']}건",  12, "FEF3C7"),
        (f"  ✓ min_tax(최저한세) 보강:       {stats['min_tax']}건",   12, "FEF3C7"),
        (f"  ✓ exclusion_codes(중복지원) 보강: {stats['exclusion_codes']}건", 12, "FEF3C7"),
        (f"  ✓ deemed_depr(감가상각의제) 추정: {stats['deemed_depr']}건 ★", 12, "FED7AA"),
        ("",                                                         12, ""),
        ("⚠️ 주의사항",                                              16, "FEE2E2"),
        ("  1. 노란색 배경 = 자동 보강된 셀 (지원 데이터 기반)",      11, ""),
        ("  2. 주황색 배경 + ★ 표시 = 추정 값 (deemed_depr 컬럼)",   11, ""),
        ("  3. 셀에 마우스 올리면 근거 출처가 코멘트로 표시됨",      11, ""),
        ("  4. 검토 후 ★ 표시는 제거하고 실제 값으로 확정",          11, ""),
        ("  5. 확정 후 GitHub data/조특법.xlsx 로 업로드하면 사이트에 자동 반영됨", 11, ""),
        ("",                                                         12, ""),
        ("🔗 변환 기준 데이터:",                                     14, "DBEAFE"),
        ("  - agri_tax: 농어촌특별세_조특법감면_과세비과세_조견표.xlsx",  11, ""),
        ("  - min_tax:  조특법_132조_최저한세_조견표.xlsx",         11, ""),
        ("  - exclusion_codes: 조특법_127조_중복지원배제_조견표.xlsx", 11, ""),
        ("  - deemed_depr: 별도 자료 없음 → 공제명 패턴 기반 추정",  11, ""),
    ]
    for i, (text, size, bg) in enumerate(rows, 1):
        c = info.cell(i, 1)
        c.value = text
        c.font = Font(size=size, bold=(size >= 14))
        c.alignment = Alignment(vertical="center")
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)

    wb.save(OUT_XLSX)

    print(f"[OK] {OUT_XLSX.name} 생성 완료")
    print(f"     경로: {OUT_XLSX}")
    print()
    print("보강 결과:")
    for k, v in stats.items():
        print(f"  {k}: {v}건")


if __name__ == "__main__":
    main()
