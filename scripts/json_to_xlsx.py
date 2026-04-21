#!/usr/bin/env python3
"""JSON → 엑셀 환원기.

사용법:
    python scripts/json_to_xlsx.py                # 현행(data/) 기준 → data/복원_조특법_현행.xlsx
    python scripts/json_to_xlsx.py 2025           # data/archive/2025/ → data/복원_조특법_2025.xlsx
    python scripts/json_to_xlsx.py 2025 -o out.xlsx  # 출력 경로 지정

생성되는 엑셀 구조(복원):
    - 시트 '혜택내용' : 공제목록 JSON (id~carry_over 11개 컬럼)
    - 시트 '개정이력' : 개정이력 JSON (있으면)
    - 시트 '농특세_통합DB' : 조문상세 JSON → 요약 테이블
    - 시트 '127조_중복배제매트릭스' : 중복배제 페어

주의: 원본 엑셀의 색상·수식·주석은 JSON 에 저장되지 않으므로 복원되지 않음.
      데이터만 완전 복원.
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"


def load_json(path: Path, default):
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def resolve_source(year: str | None) -> Path:
    if not year or year == "current" or year == "현행":
        return DATA
    p = DATA / "archive" / year
    if not p.exists():
        sys.exit(f"[오류] 아카이브 없음: {p}")
    return p


def write_deductions_sheet(ws, deductions: list[dict]):
    headers = ["id","article","title","tags","reqs","management",
               "exclusion_codes","agri_tax","min_tax","deemed_depr","carry_over"]
    ws.append(headers)
    for i, col in enumerate(headers, 1):
        ws.cell(1, i).font = Font(bold=True)
        ws.cell(1, i).fill = PatternFill("solid", fgColor="E0E7FF")
    for d in deductions:
        ws.append([
            d.get("id",""),
            d.get("article",""),
            d.get("title",""),
            ", ".join(d.get("tags",[])),
            "\n".join(d.get("reqs",[])),
            d.get("management",""),
            ", ".join(d.get("exclusion_codes",[])) or "NONE",
            d.get("agri_tax",""),
            d.get("min_tax",""),
            d.get("deemed_depr",""),
            d.get("carry_over",""),
        ])
    # 컬럼 너비
    for col, width in zip("ABCDEFGHIJK",
                          [14, 18, 38, 24, 60, 18, 14, 12, 12, 12, 14]):
        ws.column_dimensions[col].width = width
    # 요건은 줄바꿈 표시
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 5).alignment = Alignment(wrap_text=True, vertical="top")


def write_revisions_sheet(ws, revisions: list[dict]):
    headers = ["조문","공제명","개정일","시행일","변경유형","변경내용"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        ws.cell(1, i).font = Font(bold=True)
        ws.cell(1, i).fill = PatternFill("solid", fgColor="FEF3C7")
    for r in revisions:
        ws.append([
            r.get("article",""),
            r.get("title",""),
            r.get("amend_date",""),
            r.get("effect_date",""),
            r.get("change_type",""),
            r.get("content",""),
        ])
    for col, width in zip("ABCDEF", [14, 32, 14, 14, 12, 60]):
        ws.column_dimensions[col].width = width
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 6).alignment = Alignment(wrap_text=True, vertical="top")


def write_article_detail_sheet(ws, detail: dict):
    headers = ["조문키","조문표기","명칭","농특세","농특세근거","농특세율",
               "128조플래그","132조플래그","132조_적용세목"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        ws.cell(1, i).font = Font(bold=True)
        ws.cell(1, i).fill = PatternFill("solid", fgColor="DCFCE7")
    def sort_key(k):
        # '10-2.3' 같은 key 를 숫자/문자 혼합 튜플로 정렬 (문자는 항상 뒤로)
        parts = k.replace("-", ".").split(".")
        out = []
        for p in parts:
            out.append((0, int(p)) if p.isdigit() else (1, p))
        return out

    for key in sorted(detail.keys(), key=sort_key):
        d = detail[key]
        at = d.get("agri_tax") or {}
        est = d.get("estimation_128") or {}
        mt = d.get("min_tax_132") or {}
        applies = mt.get("applies") or []
        kinds = ", ".join(sorted({a.get("kind","") for a in applies if a.get("kind")}))
        ws.append([
            key,
            d.get("article_display",""),
            d.get("title",""),
            at.get("status",""),
            at.get("basis",""),
            at.get("rate",""),
            est.get("flag",""),
            mt.get("flag",""),
            kinds,
        ])
    for col, width in zip("ABCDEFGHI",
                          [10, 16, 36, 10, 16, 8, 10, 18, 16]):
        ws.column_dimensions[col].width = width


def write_matrix_sheet(ws, payload: dict):
    pairs = (payload or {}).get("pairs", [])
    headers = ["감면A_조문","A_명칭","감면B_조문","B_명칭","관계","근거"]
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        ws.cell(1, i).font = Font(bold=True)
        ws.cell(1, i).fill = PatternFill("solid", fgColor="FEE2E2")
    for p in pairs:
        ws.append([
            p.get("a_article",""),
            p.get("a_name",""),
            p.get("b_article",""),
            p.get("b_name",""),
            p.get("relation",""),
            p.get("basis",""),
        ])
    for col, width in zip("ABCDEF", [14, 28, 14, 28, 22, 12]):
        ws.column_dimensions[col].width = width


def main():
    p = argparse.ArgumentParser(description="JSON → 엑셀 환원")
    p.add_argument("year", nargs="?", default=None,
                   help="연도 (생략 시 현행). 예: 2025")
    p.add_argument("-o", "--output", default=None, help="출력 xlsx 경로")
    args = p.parse_args()

    src = resolve_source(args.year)
    label = args.year or "현행"

    deductions = load_json(src / "조특법_공제목록.json", [])
    detail     = load_json(src / "조특법_조문상세.json", {})
    matrix     = load_json(src / "조특법_중복배제매트릭스.json", {})
    revisions  = load_json(src / "조특법_개정이력.json", [])

    wb = openpyxl.Workbook()
    # 기본 시트 제거
    wb.remove(wb.active)

    write_deductions_sheet(wb.create_sheet("혜택내용"), deductions)
    write_revisions_sheet(wb.create_sheet("개정이력"), revisions)
    write_article_detail_sheet(wb.create_sheet("농특세_통합DB"), detail)
    write_matrix_sheet(wb.create_sheet("127조_중복배제매트릭스"), matrix)

    if args.output:
        out_path = Path(args.output)
    else:
        out_path = DATA / f"복원_조특법_{label}.xlsx"

    wb.save(out_path)
    print(f"[OK] 환원 완료: {out_path}")
    print(f"     소스: {src}")
    print(f"     시트: {wb.sheetnames}")


if __name__ == "__main__":
    main()
